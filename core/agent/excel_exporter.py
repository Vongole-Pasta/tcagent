
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Optional
from core.agent.state import GeneratedScenario

logger = logging.getLogger(__name__)

# Excel 스타일 상수 정의
_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

class ExcelExporter:
    @staticmethod
    def _apply_header_style(ws, row_num: int, col_count: int):
        """헤더 행에 스타일(배경색, 폰트, 테두리)을 적용합니다."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    @staticmethod
    def _apply_cell_style(ws, row_num: int, col_count: int):
        """데이터 행에 테두리와 정렬을 적용합니다."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = _THIN_BORDER
            cell.alignment = _CELL_ALIGNMENT

    @staticmethod
    def _auto_adjust_column_width(ws):
        """컬럼 너비를 내용에 맞게 자동 조절합니다. (최소 10, 최대 60)"""
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        # 한글은 2칸, 영문은 1칸으로 계산
                        char_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, char_len)
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def create_workbook(scenarios: List[GeneratedScenario], summary: Optional[str], file_path: str):
        """
        'Summary', 'VOD', '시나리오' 시트가 포함된 Excel 워크북을 생성합니다.
        전문적인 스타일링(헤더 색상, 테두리, 자동 너비 조절)이 적용됩니다.
        """
        wb = Workbook()
        
        # --- Summary 시트 ---
        if summary:
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_summary["A1"] = "테스트 전략 요약"
            ws_summary["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
            ws_summary["A2"] = summary
            ws_summary["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            ws_summary.column_dimensions["A"].width = 100
            
            ws_vod = wb.create_sheet("VOD")
        else:
            ws_vod = wb.active
            ws_vod.title = "VOD"
        
        # --- VOD 시트 ---
        headers_vod = [
            "Test Case ID", "Test Case Name", "Step No", "Description", 
            "Pre-condition", "Procedure", "Expected Result", "API Endpoint", 
            "", "", "", "", "", "", "", "", "Scenario ID"
        ]
        ws_vod.append(headers_vod)
        ExcelExporter._apply_header_style(ws_vod, 1, len(headers_vod))
        
        # 필터 자동 적용
        ws_vod.auto_filter.ref = f"A1:Q{len(scenarios) + 1}"
        
        for row_idx, scenario in enumerate(scenarios, start=2):
            row = [
                scenario.test_case_id,
                scenario.test_case_name,
                scenario.step_no,
                scenario.description,
                scenario.pre_condition,
                scenario.procedure,
                scenario.expected_result,
                scenario.api_endpoint or scenario.root_method_signature,
                "", "", "", "", "", "", "", "",
                scenario.scenario_id
            ]
            ws_vod.append(row)
            ExcelExporter._apply_cell_style(ws_vod, row_idx, len(headers_vod))

        # 틀 고정 (헤더 행 고정)
        ws_vod.freeze_panes = "A2"
        ExcelExporter._auto_adjust_column_width(ws_vod)

        # --- 시나리오 시트 ---
        ws_scenario = wb.create_sheet("시나리오")
        headers_scenario = ["Scenario ID", "Func Name", "Description"]
        ws_scenario.append(headers_scenario)
        ExcelExporter._apply_header_style(ws_scenario, 1, len(headers_scenario))
        
        seen_scenarios = set()
        row_idx = 2
        for scenario in scenarios:
            if scenario.scenario_id not in seen_scenarios:
                ws_scenario.append([
                    scenario.scenario_id,
                    scenario.api_endpoint or scenario.root_method_signature,
                    scenario.description
                ])
                ExcelExporter._apply_cell_style(ws_scenario, row_idx, len(headers_scenario))
                seen_scenarios.add(scenario.scenario_id)
                row_idx += 1
        
        ws_scenario.freeze_panes = "A2"
        ExcelExporter._auto_adjust_column_width(ws_scenario)
        
        try:
            wb.save(file_path)
            logger.info(f"Excel 파일 저장 완료: {file_path}")
        except Exception as e:
            logger.error(f"Excel 파일 저장 실패: {e}")
            raise
